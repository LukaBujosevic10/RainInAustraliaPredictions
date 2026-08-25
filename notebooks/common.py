"""
Zajednicke funkcije/konstante/import-i izvuceni IZ FinalDocument.ipynb, bez izmene
logike (izuzetak: kreirajAnalizu() sada prima/vraca `data` kao parametar umesto da
koristi `global data`, i write_log() dobija odbrambenu proveru za slucaj kad je
old_df=None - oboje je neophodno jer svaka sveska sada ima svoj nezavisan kernel,
pa vise ne dele jednu globalnu promenljivu kao u originalnoj monolitnoj svesci).

Svaka numerisana sveska u ovom folderu pocinje sa `from common import *`.
FinalDocument.ipynb je ostao NETAKNUT - ovo je samo relokacija istog koda.
"""

import os
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(_PROJECT_ROOT)
from sklearn.experimental import enable_iterative_imputer
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import math
import csv
import mlflow
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from xgboost import XGBClassifier
from sklearn.tree import DecisionTreeClassifier
from itertools import combinations
import networkx as nx
import warnings
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score,accuracy_score, precision_score, recall_score, f1_score, roc_auc_score,classification_report
from sklearn.ensemble import RandomForestRegressor,IsolationForest,RandomForestClassifier
from sklearn.neighbors import KNeighborsRegressor,NearestNeighbors
from sklearn.linear_model import LinearRegression 
import random as rnd
from typing import Dict, List, Tuple
from sklearn.impute import KNNImputer, IterativeImputer,SimpleImputer
from sklearn.preprocessing import StandardScaler,LabelEncoder
from sklearn.covariance import MinCovDet
from sklearn.feature_selection import mutual_info_classif
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from xgboost import XGBRegressor
from sklearn.neighbors import LocalOutlierFactor

try:
    import geopandas as gpd
    from shapely.geometry import Point, LineString
    import cartopy.io.shapereader as shpreader
    import pyproj
    from shapely.ops import transform, linemerge
    import os
    GEO_LIBS_AVAILABLE = True
except ImportError as e:
    print(f"[info] Geoprostorni paketi nisu dostupni ({e}); koristice se vec sacuvani"
          " OutputData/weaterAusStanice.csv umesto ponovnog racunanja.")
    GEO_LIBS_AVAILABLE = False
import re
import pandas as pd
import pytz
try:
    from timezonefinder import TimezoneFinder
    TIMEZONEFINDER_AVAILABLE = True
except ImportError as e:
    print(f"[info] timezonefinder nije dostupan ({e}); koristice se vec sacuvan"
          " OutputData/tacka_rose.csv umesto ponovnog racunanja.")
    TIMEZONEFINDER_AVAILABLE = False
from datetime import timedelta
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import RFE
import mlflow

import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
import warnings

warnings.filterwarnings('ignore')
import logging

# Postavlja MLflow logger da prikazuje samo kritične greške
logging.getLogger("mlflow").setLevel(logging.ERROR)
os.environ["MLFLOW_SKLEARN_IGNORE_WARNINGS"] = "true"
os.makedirs("logs", exist_ok=True)
os.makedirs("backups", exist_ok=True)
os.makedirs("reports", exist_ok=True)
os.makedirs("reports/claster_corelations", exist_ok=True)
os.makedirs("AnomalyDetectionResults", exist_ok=True)

with open("logs/logFile.txt", "w", encoding="utf-8") as log_file:
    print(f"", file=log_file)
old_df = None
def write_log(new_df, msg, backupName):
    global old_df
    if old_df is None:
        old_df = new_df.copy()
    with open("logs/logFile.txt", "a", encoding="utf-8") as log_file:
        print(f"--- {msg} ---", file=log_file)
        old_shape = old_df.shape
        new_shape = new_df.shape
        print(f"Dimenzije starog DF: {old_shape}", file=log_file)
        print(f"Dimenzije novog DF: {new_shape}", file=log_file)
        if old_shape != new_shape:
            print(f" Dimenzije su se promenile: {new_shape[0]-old_shape[0]} redova, {new_shape[1]-old_shape[1]} kolona\n", file=log_file)
        else:
            print(" Dimenzije su iste\n", file=log_file)

        old_cols = set(old_df.columns)
        new_cols = set(new_df.columns)

        added_cols = new_cols - old_cols
        removed_cols = old_cols - new_cols

        if added_cols:
            print(f" Nove kolone: {added_cols}", file=log_file)
        if removed_cols:
            print(f" Obrisane kolone: {removed_cols}", file=log_file)
        if not added_cols and not removed_cols:
            print(" Nema promena u kolonama\n", file=log_file)

        old_missing = old_df.isna().sum()
        new_missing = new_df.isna().sum()

        changed_missing = {}
        for col in old_cols & new_cols:
            if old_missing[col] != new_missing[col]:
                changed_missing[col] = (old_missing[col], new_missing[col])

        if changed_missing:
            print("Promene u broju nedostajućih vrednosti po kolonama:", file=log_file)
            for col, (old_val, new_val) in changed_missing.items():
                diff = new_val - old_val
                print(f" - {col}: stari={old_val}, novi={new_val}, promena={diff}", file=log_file)
        else:
            print("Broj nedostajućih vrednosti je isti u svim kolonama\n", file=log_file)
        old_df = new_df.copy()
        if backupName is not None:
            backupPath = "backups/"+backupName
        
            new_df.to_csv(backupPath, index=False)
            print(f"Backup fajl sacuvan na putanji " + backupPath, file=log_file)
        print(f"--- Kraj loga za: {msg} ---\n\n", file=log_file)
def loadData(fileName, toDataColumns = ["Date"], sortByColumns = ['Location', 'Date']) -> pd.DataFrame:
    data = pd.read_csv(fileName)
    for date in toDataColumns:
        data[date] = pd.to_datetime(data[date])
    data = data.sort_values(sortByColumns)
    return data

SPLIT_DATE = pd.Timestamp("2015-11-10")  # 80. percentil datuma na sirovom skupu

def gap_validan(df, broj_koraka):
    if broj_koraka == 0:
        return pd.Series(True, index=df.index)
    razmak = df.groupby('Location')['Date'].diff(broj_koraka).dt.days
    return razmak == broj_koraka
def calculateDiff(df, col1, col2, new_col_name):
    df[new_col_name] = (df[col1] - df[col2]).round(2)
    return df
def calculateDiffs(df):
    df = calculateDiff(df, 'MaxTemp', 'MinTemp', 'Max_Min_Temp_Diff')
    df = calculateDiff(df, 'MaxTemp', 'Temp3pm', 'Max_Temp3pm_diff')
    df = calculateDiff(df, 'MaxTemp', 'Temp3pm', 'Max_Temp3pm_diff')
    df = calculateDiff(df, 'Temp9am', 'MinTemp', 'Temp9am_Min_diff')

    df = calculateDiff(df, 'WindDir3pm_sin', 'WindDir9am_sin', 'Wind_sin_diff_3pm_9am')
    df = calculateDiff(df, 'WindDir3pm_cos', 'WindDir9am_cos', 'Wind_cos_diff_3pm_9am')
    df =calculateDiff(df, 'WindSpeed3pm', 'WindSpeed9am', 'WindSpeed_diff_3pm_9am')
    df =calculateDiff(df, 'Pressure3pm', 'Pressure9am', 'Pressure_diff_3pm_9am')
    df =calculateDiff(df, 'Humidity3pm', 'Humidity9am', 'Humidity_diff_3pm_9am')
 
    validno_juce = gap_validan(df, 1)
    df['Wind_sin_diff_9am_3pm_juce'] = (df['WindDir9am_sin'] - df.groupby('Location')['WindDir3pm_sin'].shift(1)).round(2)
    df['Wind_cos_diff_9am_3pm_juce'] = (df['WindDir9am_cos'] - df.groupby('Location')['WindDir3pm_cos'].shift(1)).round(2)
    df['WindSpeed_diff_9am_3pm_juce'] = (df['WindSpeed9am'] - df.groupby('Location')['WindSpeed3pm'].shift(1)).round(2)
    df.loc[~validno_juce, ['Wind_sin_diff_9am_3pm_juce', 'Wind_cos_diff_9am_3pm_juce', 'WindSpeed_diff_9am_3pm_juce']] = np.nan
    return df

def calculateRollingMean(df, col_name, window_size):
    novo_ime = col_name + f'_rolling_mean_{window_size}'
    df[novo_ime] = (
        df.groupby('Location')[col_name]
          .rolling(window=window_size)
          .mean()
          .reset_index(0, drop=True)
    ).round(2)
    # prozor od window_size redova mora pokrivati tacno window_size-1 dana razmaka,
    # inace prozor prelazi preko rupe u seriji i prosek nije stvarno 'poslednja N dana'
    df.loc[~gap_validan(df, window_size - 1), novo_ime] = np.nan
    return df
def calculateRollingMeans(df):
    df = calculateRollingMean(df, 'WindSpeed3pm', 3)
    df = calculateRollingMean(df, 'WindSpeed9am', 3)
    df = calculateRollingMean(df, 'Temp3pm', 3)
    return df

def calculatePreviousValue(df, col_name, brDanaPre):
    novo_ime = col_name + '_pre_'+str(brDanaPre)+"_dana"
    df[novo_ime] = df.groupby('Location')[col_name].shift(brDanaPre)
    df.loc[~gap_validan(df, brDanaPre), novo_ime] = np.nan
    return df
def calculatePreviousValues(df):
    df = calculatePreviousValue(df, 'Pressure9am', 1)
    df = calculatePreviousValue(df, 'Pressure3pm', 1)
    return df

def createVaribales(df):
    df = calculatePreviousValues(df)
    df = calculateRollingMeans(df)
    df = calculateDiffs(df)
    return df
def kreirajAnalizu(data):
    TARGET = "RainTomorrow"
    is_train_local = data['Date'] < SPLIT_DATE

    missing_pct = data.loc[is_train_local].isna().mean().mul(100).round(2)
    missing_pct.name = "Missing_%"

    data = data[is_train_local].copy()

    target_encoder = LabelEncoder()
    y = target_encoder.fit_transform(data[TARGET].astype(str))

    X = data.drop(columns=[TARGET])

    if "Date" in X.columns:
        X["Date"] = pd.to_datetime(X["Date"], errors="coerce")
        X["Month"] = X["Date"].dt.month
        X = X.drop(columns=["Date"])

    categorical_cols = X.select_dtypes(include=["object"]).columns.tolist()
    numeric_cols = [col for col in X.select_dtypes(include=[np.number]).columns.tolist() if not col.endswith("missing")]

    X_encoded = X.copy()

    if categorical_cols:
        cat_imputer = SimpleImputer(strategy="most_frequent")
        X_encoded[categorical_cols] = cat_imputer.fit_transform(X[categorical_cols])
        for col in categorical_cols:
            le = LabelEncoder()
            X_encoded[col] = le.fit_transform(X_encoded[col].astype(str))

    if numeric_cols:
        num_imputer = SimpleImputer(strategy="median")
        X_encoded[numeric_cols] = num_imputer.fit_transform(X[numeric_cols])

    feature_names = X_encoded.columns.tolist()

    rf = RandomForestClassifier(
        n_estimators=300,
        max_depth=None,
        random_state=42,
        n_jobs=-1,
        class_weight="balanced",
    )
    rf.fit(X_encoded, y)

    rf_importance = pd.Series(rf.feature_importances_, index=feature_names, name="RF_importance")

    discrete_mask = [col in categorical_cols for col in feature_names]

    mi_scores = mutual_info_classif(
        X_encoded, y, discrete_features=discrete_mask, random_state=42
    )
    mi_importance = pd.Series(mi_scores, index=feature_names, name="MutualInfo")

    summary = pd.concat(
        [missing_pct.reindex(feature_names), rf_importance, mi_importance],
        axis=1,
    )

    summary["RF_importance_norm"] = summary["RF_importance"] / summary["RF_importance"].sum()
    summary["MutualInfo_norm"] = summary["MutualInfo"] / summary["MutualInfo"].sum()

    rf_vals = summary["RF_importance_norm"].clip(lower=0)
    mi_vals = summary["MutualInfo_norm"].clip(lower=0)
    summary["Geometrijska sredina RFi MI"] = np.sqrt(rf_vals * mi_vals)

    summary = summary.sort_values("Geometrijska sredina RFi MI", ascending=False)

    print("\n=== Unakrsna matrica: značaj atributa vs. procenat nedostajućih vrednosti ===\n")
    print(summary.round(4).to_string())

    excel_path = "reports/znacaj_i_nedostajuce_vrednosti.xlsx"

    export_df = summary.reset_index().rename(columns={"index": "Atribut"})
    export_df = export_df.round(4)

    export_df.to_excel(excel_path, index=False, sheet_name="Znacaj_atributa")

    wb = load_workbook(excel_path)
    ws = wb["Znacaj_atributa"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name="Arial")

    for col_idx, col_name in enumerate(export_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font

    for col_idx, col_name in enumerate(export_df.columns, start=1):
        max_len = max(export_df[col_name].astype(str).map(len).max(), len(col_name)) + 3
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    ws.freeze_panes = "A2"
    wb.save(excel_path)

    print(f"\nExcel izveštaj sačuvan u: {excel_path}")

    plot_data = summary[["Missing_%", "RF_importance_norm", "MutualInfo_norm"]].copy()
    plot_data.columns = ["Nedostajuće (%)", "Značaj (RF, norm.)", "Značaj (MI, norm.)"]

    plot_data_scaled = (plot_data - plot_data.min()) / (plot_data.max() - plot_data.min())

    plt.figure(figsize=(8, max(6, 0.35 * len(plot_data_scaled))))
    sns.heatmap(
        plot_data_scaled,
        annot=plot_data.round(2),
        fmt="",
        cmap="YlOrRd",
        linewidths=0.5,
        cbar_kws={"label": "Relativna vrednost (0-1, skalirano po koloni)"},
    )
    plt.title("Značaj atributa za predikciju RainTomorrow\nvs. stepen nedostajućih vrednosti")
    plt.tight_layout()
    plt.savefig("graphs/matrica_znacaja_i_nedostajucih.png", dpi=200)
    plt.show()


    plt.figure(figsize=(10, 8))

    x_vals = summary["Missing_%"]
    y_vals = summary["Geometrijska sredina RFi MI"]

    plt.scatter(x_vals, y_vals, s=80, color="#C0392B", edgecolor="black", alpha=0.8, zorder=3)

    for atribut in summary.index:
        plt.annotate(
            atribut,
            xy=(x_vals[atribut], y_vals[atribut]),
            xytext=(6, 4),
            textcoords="offset points",
            fontsize=8,
        )

    plt.xlabel("Procenat nedostajućih vrednosti (%)")
    plt.ylabel("Geometrijska sredina (RF, MI) - normalizovano")
    plt.title("Značaj atributa (geometrijska sredina RF i MI)\nvs. stepen nedostajućih vrednosti")
    plt.grid(True, linestyle="--", alpha=0.4, zorder=0)
    plt.tight_layout()
    plt.savefig("graphs/scatter_znacaj_vs_nedostajuci.png", dpi=200)
    plt.show()

    plt.figure(figsize=(10, 8))

    plt.scatter(x_vals, y_vals, s=80, color="#2980B9", edgecolor="black", alpha=0.8, zorder=3)

    for atribut in summary.index:
        if 0 <= x_vals[atribut] <= 10 and 0 <= y_vals[atribut] <= 0.08:
            plt.annotate(
                atribut,
                xy=(x_vals[atribut], y_vals[atribut]),
                xytext=(6, 4),
                textcoords="offset points",
                fontsize=8,
            )

    plt.xlim(0, 8)
    plt.ylim(0, 0.065)
    plt.xlabel("Procenat nedostajućih vrednosti (%)")
    plt.ylabel("Geometrijska sredina (RF, MI) - normalizovano")
    plt.title("Zumirani prikaz: Značaj atributa vs. stepen nedostajućih vrednosti\n(X: 0-10, Y: 0-0.08)")
    plt.grid(True, linestyle="--", alpha=0.4, zorder=0)
    plt.tight_layout()
    plt.savefig("graphs/scatter_znacaj_vs_nedostajuci_zumirano.png", dpi=200)
    plt.show()
    return data


